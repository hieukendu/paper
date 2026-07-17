from __future__ import annotations

import argparse
import json
import sys
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

from vipragsent.data.schema import EMOTION_LABELS, POLARITY_LABELS, PRAGMATIC_LABELS, make_record
from vipragsent.utils.io import read_jsonl, write_jsonl

LABEL_FIELDS = [*PRAGMATIC_LABELS, "polarity", "emotion"]


def _binary(value: Any) -> int:
    if value in (True, 1, "1") or str(value).strip().upper() == "TRUE":
        return 1
    if value in (False, 0, "0") or str(value).strip().upper() == "FALSE":
        return 0
    raise ValueError(f"invalid binary label: {value!r}")


def load_workbook_records(path: Path, annotator_id: str, role: str) -> dict[str, dict]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook["annotation"]
    rows = worksheet.iter_rows(values_only=True)
    header = [str(value) for value in next(rows)]
    required = {"record_id", *LABEL_FIELDS}
    missing = required - set(header)
    if missing:
        raise ValueError(f"{path}: missing columns {sorted(missing)}")
    indexes = {name: header.index(name) for name in header}
    records: dict[str, dict] = {}
    for row_number, row in enumerate(rows, start=2):
        record_id = str(row[indexes["record_id"]] or "").strip()
        if not record_id:
            continue
        if record_id in records:
            raise ValueError(f"{path}:{row_number}: duplicate record_id {record_id}")
        labels = {label: _binary(row[indexes[label]]) for label in PRAGMATIC_LABELS}
        labels["polarity"] = str(row[indexes["polarity"]]).strip().lower()
        labels["emotion"] = str(row[indexes["emotion"]]).strip().lower()
        if labels["polarity"] not in POLARITY_LABELS:
            raise ValueError(f"{path}:{row_number}: invalid polarity {labels['polarity']!r}")
        if labels["emotion"] not in EMOTION_LABELS:
            raise ValueError(f"{path}:{row_number}: invalid emotion {labels['emotion']!r}")
        record = {"id": record_id, "labels": labels}
        if role == "adjudicator":
            record["adjudicator_id"] = annotator_id
        else:
            record["annotator_id"] = annotator_id
        note_index = indexes.get("reviewer_note")
        if note_index is not None and row[note_index] not in (None, ""):
            record["reviewer_note"] = str(row[note_index])
        record["_row"] = row_number
        for field in ("sample", "batch_id", "platform", "source_dataset", "source_path", "text_raw"):
            field_index = indexes.get(field)
            record[f"_{field}"] = row[field_index] if field_index is not None else None
        records[record_id] = record
    return records


def load_batches(path: Path) -> tuple[dict[str, dict], dict[str, list[str]]]:
    records: dict[str, dict] = {}
    batch_ids: dict[str, list[str]] = {}
    for batch_path in sorted(path.glob("*.jsonl")):
        for record in read_jsonl(batch_path):
            record_id = str(record["id"])
            if record_id in records:
                raise ValueError(f"duplicate batch record_id {record_id}")
            records[record_id] = record
            batch_id = str(record.get("batch_id") or batch_path.stem.replace("_input", ""))
            batch_ids.setdefault(batch_id, []).append(record_id)
    return records, batch_ids


def main() -> int:
    parser = argparse.ArgumentParser(description="Import the three ViPragSent annotation workbooks.")
    parser.add_argument("--reviewer-1", type=Path, default=ROOT / "Duy_Duc_synced-1.xlsx")
    parser.add_argument("--reviewer-2", type=Path, default=ROOT / "Nhat_Khang_synced-1.xlsx")
    parser.add_argument("--adjudicator", type=Path, default=ROOT / "Quynh_Nhu_synced-1.xlsx")
    args = parser.parse_args()

    old_batches, _ = load_batches(ROOT / "data" / "annotation" / "batches")
    reviewer_1 = load_workbook_records(args.reviewer_1, "duy_duc", "reviewer")
    reviewer_2 = load_workbook_records(args.reviewer_2, "nhat_khang", "reviewer")
    adjudicator = load_workbook_records(args.adjudicator, "quynh_nhu", "adjudicator")
    expected = set(adjudicator)
    if len(expected) != 12000:
        raise ValueError(f"adjudicator workbook must contain 12000 unique IDs, got {len(expected)}")
    shared_review_ids = expected & set(reviewer_1) & set(reviewer_2)
    coverage = {
        "reviewer_1": {"records_in_workbook": len(reviewer_1), "canonical_records": len(expected & set(reviewer_1))},
        "reviewer_2": {"records_in_workbook": len(reviewer_2), "canonical_records": len(expected & set(reviewer_2))},
        "adjudicator": {"records": len(adjudicator), "canonical_records": len(expected)},
        "shared_reviewer_records": len(shared_review_ids),
        "adjudicator_only_replacement_records": len(expected - shared_review_ids),
        "replaced_old_batch_records": len(set(old_batches) - expected),
    }

    # Quynh Nhu's workbook is canonical: it intentionally replaces 2,000 old
    # ViSoBERT samples with generated candidates while retaining row/batch slots.
    batch_ids: dict[str, list[str]] = {}
    canonical_batches: dict[str, list[dict]] = {}
    for record in sorted(adjudicator.values(), key=lambda value: int(value["_row"])):
        batch_id = str(record["_batch_id"])
        text = str(record["_sample"] or record["_text_raw"] or "").strip()
        dataset = str(record["_source_dataset"] or "unknown")
        platform = str(record["_platform"] or "unknown")
        base = make_record(
            text,
            dataset=dataset,
            platform=platform,
            source_path=str(record["_source_path"] or "") or None,
            record_id=record["id"],
            split="unassigned",
            label_status="unlabeled",
            created_by=f"scripts/12_import_annotation_workbooks.py:{args.adjudicator.name}",
            pii_cleaned=True,
        )
        base["batch_id"] = batch_id
        base["annotation"]["batch_id"] = batch_id
        batch_ids.setdefault(batch_id, []).append(record["id"])
        canonical_batches.setdefault(batch_id, []).append(base)
    batches_dir = ROOT / "data" / "annotation" / "batches"
    for old in batches_dir.glob("*.jsonl"):
        old.unlink()
    for batch_id, records in sorted(canonical_batches.items()):
        write_jsonl(batches_dir / f"{batch_id}_input.jsonl", records)

    directories = {
        "reviewer_1": ROOT / "data" / "annotation" / "reviewer_01",
        "reviewer_2": ROOT / "data" / "annotation" / "reviewer_02",
        "adjudicator": ROOT / "data" / "annotation" / "adjudicated",
    }
    sources = {"reviewer_1": reviewer_1, "reviewer_2": reviewer_2, "adjudicator": adjudicator}
    for directory in directories.values():
        directory.mkdir(parents=True, exist_ok=True)
        for old in directory.glob("*.jsonl"):
            old.unlink()
    for batch_id, ids in sorted(batch_ids.items()):
        for name, records in sources.items():
            selected = []
            for record_id in ids:
                if record_id not in records:
                    continue
                selected.append({key: value for key, value in records[record_id].items() if not key.startswith("_")})
            if selected:
                write_jsonl(directories[name] / f"{batch_id}.jsonl", selected)

    disagreements = []
    field_counts: Counter[str] = Counter()
    agreement_count = 0
    for record_id in sorted(shared_review_ids):
        left = reviewer_1[record_id]
        right = reviewer_2[record_id]
        fields = [field for field in LABEL_FIELDS if left["labels"][field] != right["labels"][field]]
        if not fields:
            agreement_count += 1
            continue
        field_counts.update(fields)
        disagreements.append({
            "id": record_id,
            "reviewer_1": left,
            "reviewer_2": right,
            "disagreement_fields": fields,
            "needs_adjudication": True,
            "adjudication": adjudicator[record_id],
            "resolved": True,
        })
    disagreement_path = ROOT / "data" / "annotation" / "disagreements" / "resolved_disagreements.jsonl"
    write_jsonl(disagreement_path, disagreements)
    report = {
        "status": "ok",
        "base_records": len(expected),
        "reviewer_comparable_records": len(shared_review_ids),
        "adjudicator_only_replacement_records": len(expected - shared_review_ids),
        "exact_agreements": agreement_count,
        "records_with_disagreement": len(disagreements),
        "disagreement_field_occurrences": dict(sorted(field_counts.items())),
        "coverage": coverage,
        "mapping": {
            args.reviewer_1.name: "reviewer_01 / duy_duc",
            args.reviewer_2.name: "reviewer_02 / nhat_khang",
            args.adjudicator.name: "adjudicated / quynh_nhu (final gold)",
        },
        "disagreements_output": disagreement_path.relative_to(ROOT).as_posix(),
    }
    report_path = ROOT / "data" / "manifest" / "annotation_import_report.json"
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
